from __future__ import annotations

import json
import re
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agent_compliance.core.config import detect_paths
from agent_compliance.agents.compliance_review.pipelines.procurement_stage_router import route_procurement_stage
from agent_compliance.agents.compliance_review.pipelines.rewrite_generator import (
    ACTION_DIRECT,
    ACTION_JUSTIFY,
    ACTION_PREFIXES,
    ACTION_REVIEW,
    ACTION_SOFTEN,
    determine_suggested_action,
)
from agent_compliance.core.schemas import Finding, ReviewResult


def export_review_bytes(
    review: ReviewResult,
    *,
    export_format: str,
    mode: str,
    document_payload: dict[str, Any] | None = None,
) -> tuple[bytes, str, str]:
    normalized_format = export_format.lower()
    normalized_mode = "summary" if mode == "summary" else "full"
    if normalized_format == "json":
        payload = build_export_payload(review, mode=normalized_mode, document_payload=document_payload)
        content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        return content, "application/json; charset=utf-8", build_export_filename(review.document_name, normalized_mode, "json")
    if normalized_format == "markdown":
        content = render_export_markdown(review, mode=normalized_mode, document_payload=document_payload).encode("utf-8")
        return content, "text/markdown; charset=utf-8", build_export_filename(review.document_name, normalized_mode, "md")
    if normalized_format == "xlsx":
        content = render_export_xlsx(review, mode=normalized_mode, document_payload=document_payload)
        return content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", build_export_filename(review.document_name, normalized_mode, "xlsx")
    raise ValueError(f"不支持的导出格式：{export_format}")


def write_export_output(
    review: ReviewResult,
    *,
    export_format: str,
    mode: str,
    document_payload: dict[str, Any] | None = None,
) -> Path:
    content, _content_type, filename = export_review_bytes(
        review,
        export_format=export_format,
        mode=mode,
        document_payload=document_payload,
    )
    paths = detect_paths()
    export_root = paths.generated_root / "exports" / datetime.now().strftime("%Y-%m-%d")
    export_root.mkdir(parents=True, exist_ok=True)
    target = export_root / filename
    target.write_bytes(content)
    return target


def build_export_payload(
    review: ReviewResult,
    *,
    mode: str,
    document_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    findings = _pick_findings(review.findings, mode)
    stage_profile = route_procurement_stage(findings=review.findings)
    summary = {
        "overall_risk_summary": review.overall_risk_summary,
        "finding_count": len(findings),
        "high_risk_count": sum(1 for item in findings if item.risk_level == "high"),
        "medium_risk_count": sum(1 for item in findings if item.risk_level == "medium"),
        "low_risk_count": sum(1 for item in findings if item.risk_level == "low"),
        "procurement_stage_name": stage_profile.stage_name,
        "procurement_stage_goal": stage_profile.stage_goal,
        "release_recommendation": _release_recommendation(review, findings=findings),
    }
    document = {
        "document_name": review.document_name,
        "source_path": (document_payload or {}).get("source_path"),
        "normalized_text_path": (document_payload or {}).get("normalized_text_path"),
        "review_scope": review.review_scope,
        "jurisdiction": review.jurisdiction,
        "primary_catalog_name": (document_payload or {}).get("primary_catalog_name"),
        "secondary_catalog_names": (document_payload or {}).get("secondary_catalog_names", []),
        "is_mixed_scope": bool((document_payload or {}).get("is_mixed_scope")),
        "catalog_confidence": (document_payload or {}).get("catalog_confidence"),
    }
    return {
        "document": document,
        "review_summary": summary,
        "export_meta": {
            "export_format": "json",
            "export_mode": mode,
            "export_intent": "采购人改稿与发布前复核优先",
            "export_timestamp": datetime.now().isoformat(timespec="seconds"),
            "generated_by": "agent_compliance.review_export",
        },
        "findings": [
            _serialize_finding(
                item,
                mode=mode,
                handling_order=index,
                stage_profile=stage_profile,
            )
            for index, item in enumerate(findings, start=1)
        ],
        "items_for_human_review": review.items_for_human_review,
        "review_limitations": review.review_limitations,
    }


def render_export_markdown(
    review: ReviewResult,
    *,
    mode: str,
    document_payload: dict[str, Any] | None = None,
) -> str:
    findings = _pick_findings(review.findings, mode)
    stage_profile = route_procurement_stage(findings=review.findings)
    lines = [
        f"# {review.document_name} 采购需求合规性检查智能体审查结果导出",
        "",
        "## 文件信息",
        "",
        f"- 审查范围：`{review.review_scope}`",
        f"- 审查时间：`{review.review_timestamp}`",
        f"- 审查阶段：`{stage_profile.stage_name}`",
        f"- 导出模式：`{'主问题版' if mode == 'summary' else '完整明细版'}`",
        "- 导出意图：`采购人改稿与发布前复核优先`",
        f"- 发布建议：`{_release_recommendation(review, findings=findings)}`",
    ]
    if document_payload:
        lines.append(f"- 原文件：`{document_payload.get('source_path', '')}`")
        if document_payload.get("primary_catalog_name"):
            lines.append(f"- 主品目：`{document_payload.get('primary_catalog_name')}`")
        if document_payload.get("secondary_catalog_names"):
            lines.append(
                f"- 次品目：`{'、'.join(document_payload.get('secondary_catalog_names') or [])}`"
            )
        if document_payload.get("is_mixed_scope"):
            lines.append("- 混合采购：`是`")
    lines.extend(
        [
            "",
            "## 风险摘要",
            "",
            f"- 风险摘要：{review.overall_risk_summary}",
            f"- 问题数量：`{len(findings)}`",
            "",
            "## 问题清单",
            "",
        ]
    )

    for finding in findings:
        lines.extend(
            [
                f"### {finding.finding_id} {finding.problem_title}",
                f"- 章节：`{_chapter_group(finding)}`",
                f"- 风险等级：`{finding.risk_level}`",
                f"- 置信度：`{finding.confidence}`",
                f"- 合规判断：`{finding.compliance_judgment}`",
                f"- 位置：`{_full_location(finding)}`",
                f"- 问题类型：`{finding.issue_type}`",
                f"- 代表性证据：`{finding.source_text}`" if mode == "summary" else f"- 原文摘录：`{finding.source_text}`",
                f"- 风险说明：{finding.why_it_is_risky}",
                f"- 主依据/辅依据：{_authority_layers_text(finding)}",
                f"- 条文要点：{finding.authority_key_points or '暂无'}",
                f"- 适用逻辑：{finding.applicability_logic or finding.human_review_reason or '暂无'}",
                f"- 修改建议：{finding.rewrite_suggestion}",
                "",
            ]
        )

    if review.items_for_human_review:
        lines.extend(["## 需人工复核", ""])
        for item in review.items_for_human_review:
            lines.append(f"- {item}")
        lines.append("")

    return "\n".join(lines)


def build_export_filename(document_name: str, mode: str, extension: str) -> str:
    stem = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", document_name).strip("._") or "review-export"
    stem = re.sub(r"\.(docx|pdf|txt|md|json)$", "", stem, flags=re.IGNORECASE)
    suffix = "summary" if mode == "summary" else "full"
    return f"{stem}-{suffix}.{extension}"


def render_export_xlsx(review: ReviewResult, *, mode: str, document_payload: dict[str, Any] | None = None) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "审查摘要"
    _write_summary_sheet(summary_sheet, review, mode=mode, document_payload=document_payload)
    detail_sheet = workbook.create_sheet("主问题" if mode == "summary" else "完整明细")
    rows = build_excel_rows(review, mode=mode, document_payload=document_payload)
    for row in rows:
        detail_sheet.append(row)
    _style_detail_sheet(detail_sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_excel_rows(review: ReviewResult, *, mode: str, document_payload: dict[str, Any] | None = None) -> list[list[Any]]:
    findings = _pick_findings(review.findings, mode)
    stage_profile = route_procurement_stage(findings=review.findings)
    header = [
        "处理顺序",
        "是否主问题",
        "处理建议",
        "问题标题",
        "章节",
        "风险等级",
        "置信度",
        "合规判断",
        "位置",
        "页码提示",
        "原文摘录" if mode == "full" else "代表性证据",
        "风险说明",
        "主依据/辅依据",
        "条文要点",
        "主依据",
        "辅依据",
        "适用逻辑",
        "修改建议",
        "是否需复核",
        "复核原因",
        "问题类型",
    ]
    rows: list[list[Any]] = [header]
    for finding in findings:
        action = determine_suggested_action(finding, stage_profile=stage_profile)
        rows.append(
            [
                len(rows),
                "是" if _is_main_issue(finding) else "否",
                _action_label(action),
                finding.problem_title,
                _chapter_group(finding),
                finding.risk_level,
                finding.confidence,
                finding.compliance_judgment,
                _full_location(finding),
                finding.page_hint or "",
                finding.source_text,
                finding.why_it_is_risky,
                _authority_layers_text(finding),
                finding.authority_key_points or "",
                finding.primary_authority or "",
                "；".join(finding.secondary_authorities or []),
                finding.applicability_logic or "",
                finding.rewrite_suggestion,
                "是" if finding.needs_human_review else "否",
                finding.human_review_reason or "",
                finding.issue_type,
            ]
        )
    return rows


def _write_summary_sheet(sheet, review: ReviewResult, *, mode: str, document_payload: dict[str, Any] | None = None) -> None:
    findings = _pick_findings(review.findings, mode)
    stage_profile = route_procurement_stage(findings=review.findings)
    summary_rows = [
        ["文档名称", review.document_name],
        ["审查范围", review.review_scope],
        ["审查时间", review.review_timestamp],
        ["审查阶段", stage_profile.stage_name],
        ["发布建议", _release_recommendation(review, findings=findings)],
        ["主品目", (document_payload or {}).get("primary_catalog_name") or "未识别"],
        [
            "次品目",
            "；".join((document_payload or {}).get("secondary_catalog_names") or []) or "无",
        ],
        ["是否混合采购", "是" if (document_payload or {}).get("is_mixed_scope") else "否"],
        ["导出模式", "主问题版" if mode == "summary" else "完整明细版"],
        ["导出意图", "采购人改稿与发布前复核优先"],
        ["风险摘要", review.overall_risk_summary],
        ["问题数量", len(findings)],
        ["高风险数量", sum(1 for item in findings if item.risk_level == "high")],
        ["中风险数量", sum(1 for item in findings if item.risk_level == "medium")],
        ["低风险数量", sum(1 for item in findings if item.risk_level == "low")],
        ["需人工复核", "；".join(review.items_for_human_review) if review.items_for_human_review else "无"],
    ]
    for row in summary_rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[0].fill = header_fill
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 84


def _style_detail_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    high_fill = PatternFill("solid", fgColor="FBE4D5")
    medium_fill = PatternFill("solid", fgColor="FFF2CC")
    low_fill = PatternFill("solid", fgColor="E2F0D9")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    risk_col_idx = 3
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        risk = str(row[risk_col_idx - 1].value or "")
        fill = None
        if risk == "high":
            fill = high_fill
        elif risk == "medium":
            fill = medium_fill
        elif risk == "low":
            fill = low_fill
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill is not None:
                cell.fill = fill

    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 48)


def _pick_findings(findings: list[Finding], mode: str) -> list[Finding]:
    if mode != "summary":
        return findings
    return [item for item in findings if _is_main_issue(item)]


def _is_main_issue(finding: Finding) -> bool:
    return finding.finding_origin == "analyzer" or (
        finding.finding_origin == "llm_added" and bool(re.search(r"章节|主问题", finding.problem_title or ""))
    )


def _chapter_group(finding: Finding) -> str:
    text = " ".join(
        part for part in [finding.problem_title, finding.section_path, finding.source_section] if part
    )
    if re.search(r"资格|申请人的资格要求|准入门槛", text):
        return "资格"
    if re.search(r"评分|评标信息|演示|品牌档次|认证评分|商务评分", text):
        return "评分"
    if re.search(r"技术|标准|检测报告|证明材料", text):
        return "技术"
    return "商务/验收"


def _full_location(finding: Finding) -> str:
    parts: list[str] = []
    if finding.section_path:
        parts.append(finding.section_path)
    elif finding.source_section:
        parts.append(finding.source_section)
    if finding.table_or_item_label:
        parts.append(finding.table_or_item_label)
    if finding.page_hint:
        parts.append(finding.page_hint)
    if finding.text_line_start and finding.text_line_end:
        if finding.text_line_start == finding.text_line_end:
            parts.append(f"行 {finding.text_line_start}")
        else:
            parts.append(f"行 {finding.text_line_start}-{finding.text_line_end}")
    return " | ".join(part for part in parts if part) or "未定位"


def _authority_layers_text(finding: Finding) -> str:
    parts: list[str] = []
    if finding.primary_authority:
        parts.append(f"主依据：{finding.primary_authority}")
    if finding.secondary_authorities:
        parts.append(f"辅依据：{'；'.join(finding.secondary_authorities)}")
    if not parts and finding.legal_or_policy_basis:
        return finding.legal_or_policy_basis
    return " ｜ ".join(parts) if parts else "暂无"


def _serialize_finding(
    finding: Finding,
    *,
    mode: str,
    handling_order: int | None = None,
    stage_profile=None,
) -> dict[str, Any]:
    base = finding.to_dict()
    action = determine_suggested_action(
        finding,
        stage_profile=stage_profile or route_procurement_stage(findings=[finding]),
    )
    if mode == "summary":
        return {
            "finding_id": base["finding_id"],
            "handling_order": handling_order,
            "is_main_issue": _is_main_issue(finding),
            "processing_recommendation": _action_label(action),
            "problem_title": base["problem_title"],
            "chapter_group": _chapter_group(finding),
            "risk_level": base["risk_level"],
            "confidence": base["confidence"],
            "compliance_judgment": base["compliance_judgment"],
            "source_section": base["source_section"],
            "section_path": base["section_path"],
            "table_or_item_label": base["table_or_item_label"],
            "page_hint": base["page_hint"],
            "text_line_start": base["text_line_start"],
            "text_line_end": base["text_line_end"],
            "representative_evidence": base["source_text"],
            "why_it_is_risky": base["why_it_is_risky"],
            "legal_or_policy_basis": base["legal_or_policy_basis"],
            "primary_authority": base.get("primary_authority"),
            "secondary_authorities": base.get("secondary_authorities"),
            "authority_key_points": base.get("authority_key_points"),
            "applicability_logic": base.get("applicability_logic"),
            "rewrite_suggestion": base["rewrite_suggestion"],
            "needs_human_review": base["needs_human_review"],
            "human_review_reason": base["human_review_reason"],
            "issue_type": base["issue_type"],
            "finding_origin": base.get("finding_origin", "rule"),
        }
    base["handling_order"] = handling_order
    base["is_main_issue"] = _is_main_issue(finding)
    base["processing_recommendation"] = _action_label(action)
    return base


def _action_label(action: str) -> str:
    if action == ACTION_DIRECT:
        return "建议直接修改"
    if action == ACTION_SOFTEN:
        return "建议弱化表述"
    if action == ACTION_JUSTIFY:
        return "建议补充必要性论证"
    if action == ACTION_REVIEW:
        return "建议采购/法务复核"
    prefix = ACTION_PREFIXES.get(action, "")
    return prefix.rstrip("：") or action


def _release_recommendation(review: ReviewResult, *, findings: list[Finding] | None = None) -> str:
    findings = findings or review.findings
    stage_profile = route_procurement_stage(findings=review.findings)
    actions = [determine_suggested_action(item, stage_profile=stage_profile) for item in findings]
    has_high_risk = any(item.risk_level == "high" for item in findings)
    has_justify_or_review = any(action in {ACTION_JUSTIFY, ACTION_REVIEW} for action in actions)
    if has_high_risk:
        return "建议先修改后再发布"
    if has_justify_or_review or review.items_for_human_review:
        return "建议补充论证或复核后发布"
    return "建议完成常规复核后发布"
