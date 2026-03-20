from __future__ import annotations

import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import tests._bootstrap  # noqa: F401

from agent_compliance.agents.compliance_review.pipelines.review_export import export_review_bytes, write_export_output
from agent_compliance.core.schemas import Finding, ReviewResult


class ReviewExportTest(unittest.TestCase):
    def _build_review(self) -> ReviewResult:
        main_finding = Finding(
            finding_id="F-001",
            document_name="sample.docx",
            problem_title="多个方案评分项大量使用主观分档且缺少量化锚点",
            page_hint="第12页",
            clause_id="评分项A",
            source_section="评标信息",
            section_path="第二章-评标信息-技术部分",
            table_or_item_label="技术部分评分PT",
            text_line_start=100,
            text_line_end=104,
            source_text="优得满分，良得60%，中得30%，差不得分。",
            issue_type="scoring_content_mismatch",
            risk_level="high",
            severity_score=3,
            confidence="high",
            compliance_judgment="likely_non_compliant",
            why_it_is_risky="主观分档过重。",
            impact_on_competition_or_performance="压缩客观可比性。",
            legal_or_policy_basis="主依据：政府采购需求管理办法",
            rewrite_suggestion="改为量化评分。",
            needs_human_review=False,
            human_review_reason=None,
            primary_authority="政府采购需求管理办法",
            secondary_authorities=["政府采购法实施条例"],
            authority_key_points="第十四条：采购人应当通过设定评审规则落实政府采购政策功能",
            applicability_logic="评分因素应与评审目标一致。",
            finding_origin="analyzer",
        )
        detail_finding = Finding(
            finding_id="F-002",
            document_name="sample.docx",
            problem_title="检测报告要求过严",
            page_hint="第15页",
            clause_id="3.2",
            source_section="技术要求",
            section_path="第三章-技术要求",
            table_or_item_label=None,
            text_line_start=130,
            text_line_end=133,
            source_text="需提供指定机构出具检测报告。",
            issue_type="technical_justification_needed",
            risk_level="medium",
            severity_score=2,
            confidence="medium",
            compliance_judgment="needs_human_review",
            why_it_is_risky="证明形式可能过严。",
            impact_on_competition_or_performance="增加证明门槛。",
            legal_or_policy_basis="主依据：政府采购需求管理办法",
            rewrite_suggestion="允许等效证明。",
            needs_human_review=True,
            human_review_reason="需结合品目论证。",
            primary_authority="政府采购需求管理办法",
            secondary_authorities=["财政部办公厅关于做好政府采购框架协议采购工作有关问题的通知"],
            authority_key_points="第三十一条：重点审查包括非歧视性、竞争性和履约风险等内容",
            applicability_logic="技术证明材料应与履约验证需要相匹配。",
            finding_origin="rule",
        )
        return ReviewResult(
            document_name="sample.docx",
            review_scope="政府采购需求合规性审查",
            jurisdiction=None,
            review_timestamp="2026-03-18T12:00:00+08:00",
            overall_risk_summary="评分结构失衡，技术证明要求偏严。",
            findings=[main_finding, detail_finding],
            items_for_human_review=["检测报告证明形式需结合品目复核"],
            review_limitations=[],
        )

    def test_summary_json_export_keeps_only_main_issues(self) -> None:
        content, content_type, filename = export_review_bytes(
            self._build_review(),
            export_format="json",
            mode="summary",
            document_payload={
                "source_path": "/tmp/sample.docx",
                "primary_catalog_name": "信息化平台及系统运维",
                "secondary_catalog_names": ["设备供货并安装调试"],
                "is_mixed_scope": True,
                "catalog_confidence": 0.91,
            },
        )
        payload = json.loads(content.decode("utf-8"))
        self.assertEqual(content_type, "application/json; charset=utf-8")
        self.assertTrue(filename.endswith("-summary.json"))
        self.assertEqual(len(payload["findings"]), 1)
        self.assertEqual(payload["review_summary"]["procurement_stage_name"], "采购需求形成与发布前审查")
        self.assertEqual(payload["export_meta"]["export_intent"], "采购人改稿与发布前复核优先")
        self.assertEqual(payload["review_summary"]["release_recommendation"], "建议先修改后再发布")
        self.assertEqual(payload["document"]["primary_catalog_name"], "信息化平台及系统运维")
        self.assertTrue(payload["document"]["is_mixed_scope"])
        self.assertEqual(payload["findings"][0]["handling_order"], 1)
        self.assertTrue(payload["findings"][0]["is_main_issue"])
        self.assertEqual(payload["findings"][0]["processing_recommendation"], "建议弱化表述")
        self.assertEqual(payload["findings"][0]["chapter_group"], "评分")
        self.assertEqual(payload["findings"][0]["primary_authority"], "政府采购需求管理办法")
        self.assertIn("第十四条", payload["findings"][0]["authority_key_points"])

    def test_full_markdown_export_contains_detailed_fields(self) -> None:
        content, content_type, filename = export_review_bytes(
            self._build_review(),
            export_format="markdown",
            mode="full",
            document_payload={"source_path": "/tmp/sample.docx"},
        )
        markdown = content.decode("utf-8")
        self.assertEqual(content_type, "text/markdown; charset=utf-8")
        self.assertTrue(filename.endswith("-full.md"))
        self.assertIn("# sample.docx 采购需求合规性检查智能体审查结果导出", markdown)
        self.assertIn("审查阶段：`采购需求形成与发布前审查`", markdown)
        self.assertIn("导出意图：`采购人改稿与发布前复核优先`", markdown)
        self.assertIn("F-001 多个方案评分项大量使用主观分档且缺少量化锚点", markdown)
        self.assertIn("主依据/辅依据：主依据：政府采购需求管理办法", markdown)
        self.assertIn("条文要点：第三十一条：重点审查包括非歧视性、竞争性和履约风险等内容", markdown)
        self.assertIn("适用逻辑：技术证明材料应与履约验证需要相匹配。", markdown)

    def test_summary_xlsx_export_contains_header_and_main_issue(self) -> None:
        content, content_type, filename = export_review_bytes(
            self._build_review(),
            export_format="xlsx",
            mode="summary",
            document_payload={
                "primary_catalog_name": "信息化平台及系统运维",
                "secondary_catalog_names": ["设备供货并安装调试"],
                "is_mixed_scope": True,
            },
        )
        self.assertEqual(content_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(filename.endswith("-summary.xlsx"))
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(workbook.sheetnames[0], "审查摘要")
        summary_sheet = workbook["审查摘要"]
        detail_sheet = workbook["主问题"]
        self.assertEqual(summary_sheet["A1"].value, "文档名称")
        self.assertEqual(summary_sheet["A4"].value, "审查阶段")
        self.assertEqual(summary_sheet["B4"].value, "采购需求形成与发布前审查")
        self.assertEqual(summary_sheet["A5"].value, "发布建议")
        self.assertEqual(summary_sheet["B6"].value, "信息化平台及系统运维")
        self.assertEqual(detail_sheet["A1"].value, "处理顺序")
        self.assertEqual(detail_sheet.max_row, 2)
        self.assertEqual(detail_sheet.freeze_panes, "A2")
        self.assertIsNotNone(detail_sheet.auto_filter.ref)
        self.assertEqual(detail_sheet["B2"].value, "是")
        self.assertEqual(detail_sheet["C2"].value, "建议弱化表述")
        self.assertIn("主观分档", str(detail_sheet["D2"].value))
        self.assertIn("第十四条", str(detail_sheet["N2"].value))

    def test_write_export_output_persists_file_under_generated_exports(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(tmpdir)
            generated_root = repo_root / "docs" / "generated"
            generated_root.mkdir(parents=True, exist_ok=True)
            app_paths = type(
                "AppPathsLike",
                (),
                {
                    "generated_root": generated_root,
                },
            )()
            with patch("agent_compliance.agents.compliance_review.pipelines.review_export.detect_paths", return_value=app_paths):
                path = write_export_output(
                    self._build_review(),
                    export_format="json",
                    mode="summary",
                    document_payload={"source_path": "/tmp/sample.docx"},
                )
            self.assertTrue(path.exists())
            self.assertIn("/docs/generated/exports/", str(path))
            payload = json.loads(path.read_text(encoding="utf-8"))
            self.assertEqual(len(payload["findings"]), 1)


if __name__ == "__main__":
    unittest.main()
